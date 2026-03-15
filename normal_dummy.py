"""
Script untuk generate data dummy RDKK dan SIVERVAL
di mana SEMUA petani akan di-label NORMAL oleh assign_labels().

Syarat NORMAL (dari labeling.py):
  1. tebus == diajukan persis (selisih = 0)
  2. kios tebus == kios RDKK
  3. tidak ada SP36 / organik cair
  4. tidak ada tebus tanpa pengajuan (rasio != 999)

File output: dummy_data/data_rdkk_normal.xlsx
             dummy_data/data_siverval_normal.xlsx
"""

import pandas as pd
import numpy as np
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# Jumlah petani
N_PETANI = 100
np.random.seed(99)

os.makedirs('dummy_data', exist_ok=True)

# =====================================================
# MASTER DATA
# =====================================================
kios_data = {
    'K001': 'Toko Tani Makmur',
    'K002': 'Kios Subur Jaya',
    'K003': 'Tani Sejahtera',
    'K004': 'Pupuk Berkah',
    'K005': 'Sarana Tani Mandiri',
}
kode_kios_list = list(kios_data.keys())

nama_depan = [
    'Budi', 'Siti', 'Ahmad', 'Dewi', 'Agus', 'Sri', 'Hadi', 'Rina',
    'Joko', 'Ani', 'Wahyu', 'Lestari', 'Bambang', 'Nurma', 'Eko',
    'Yanti', 'Dani', 'Wati', 'Rudi', 'Mega',
]
nama_belakang = [
    'Santoso', 'Wijaya', 'Hartono', 'Susanto', 'Rahayu', 'Pratama',
    'Saputra', 'Handayani', 'Kusuma', 'Hidayat', 'Lestari', 'Cahyono',
    'Setiawan', 'Purnama', 'Nugroho', 'Wulandari', 'Suryadi', 'Utami',
]
poktan_list = [
    'Sri Rejeki', 'Tani Mulyo', 'Karya Tani', 'Maju Jaya',
    'Harapan Baru', 'Mekar Sari', 'Subur Makmur', 'Berkah Tani',
]
desa_list = [
    'Mancak', 'Karangsari', 'Banjarsari', 'Sidodadi', 'Cikeusal',
    'Cilowong', 'Drangong', 'Kalang Anyar', 'Cipocok Jaya',
]
kabupaten_list = ['Kota Serang', 'Kab. Serang', 'Kab. Pandeglang']
kecamatan_list = ['Walantaka', 'Curug', 'Cipocok Jaya', 'Kasemen', 'Taktakan']
komoditas_list = ['Padi', 'Jagung', 'Kedelai', 'Cabai', 'Bawang Merah']

# =====================================================
# 1. GENERATE DATA RDKK
# =====================================================
niks = [
    f'35{"".join([str(np.random.randint(0, 10)) for _ in range(14)])}'
    for _ in range(N_PETANI)
]
nama_petani = [
    f'{np.random.choice(nama_depan)} {np.random.choice(nama_belakang)}'
    for _ in range(N_PETANI)
]
kode_kios_rdkk = [np.random.choice(kode_kios_list) for _ in range(N_PETANI)]

# Pupuk per MT — nilai integer bersih (tidak ada float agar selisih = 0 persis)
rdkk = {
    'Nama Penyuluh':       [f'Penyuluh {chr(65 + i % 10)}' for i in range(N_PETANI)],
    'Kode Kios Pengecer':  kode_kios_rdkk,
    'Nama Kios Pengecer':  [kios_data[k] for k in kode_kios_rdkk],
    'Gapoktan':            [None] * N_PETANI,
    'Nama Poktan':         [np.random.choice(poktan_list) for _ in range(N_PETANI)],
    'Nama Petani':         nama_petani,
    'KTP':                 niks,
    'Alamat': [
        f'Desa {np.random.choice(desa_list)} RT {np.random.randint(1,10):02d}/{np.random.randint(1,10):02d}'
        for _ in range(N_PETANI)
    ],
    'Subsektor': ['Tanaman Pangan'] * N_PETANI,

    # MT1 — semua petani aktif di MT1
    'Komoditas MT1':              [np.random.choice(komoditas_list) for _ in range(N_PETANI)],
    'Luas Lahan (Ha) MT1':        list(np.round(np.random.uniform(0.25, 3.0, N_PETANI), 2)),
    'Pupuk Urea (Kg) MT1':        list(np.random.choice([50, 100, 150, 200, 250, 300], N_PETANI).astype(int)),
    'Pupuk NPK (Kg) MT1':         list(np.random.choice([50, 100, 150, 200, 250], N_PETANI).astype(int)),
    'Pupuk NPK Formula (Kg) MT1': list(np.random.choice([0, 50, 100, 150], N_PETANI).astype(int)),
    'Pupuk Organik (Kg) MT1':     list(np.random.choice([0, 100, 200, 500], N_PETANI).astype(int)),
    'Pupuk ZA (Kg) MT1':          list(np.random.choice([0, 50, 100, 150], N_PETANI).astype(int)),

    # MT2 — 70% petani aktif di MT2
    'Komoditas MT2': [np.random.choice(komoditas_list) if np.random.random() > 0.3 else '' for _ in range(N_PETANI)],
    'Luas Lahan (Ha) MT2': [round(np.random.uniform(0.25, 2.0), 2) if np.random.random() > 0.3 else 0 for _ in range(N_PETANI)],
    'Pupuk Urea (Kg) MT2':        [int(np.random.choice([50, 100, 150, 200])) if np.random.random() > 0.3 else 0 for _ in range(N_PETANI)],
    'Pupuk NPK (Kg) MT2':         [int(np.random.choice([50, 100, 150]))      if np.random.random() > 0.3 else 0 for _ in range(N_PETANI)],
    'Pupuk NPK Formula (Kg) MT2': [int(np.random.choice([0, 50, 100]))        if np.random.random() > 0.3 else 0 for _ in range(N_PETANI)],
    'Pupuk Organik (Kg) MT2':     [int(np.random.choice([0, 100, 200]))       if np.random.random() > 0.3 else 0 for _ in range(N_PETANI)],
    'Pupuk ZA (Kg) MT2':          [int(np.random.choice([0, 50, 100]))        if np.random.random() > 0.3 else 0 for _ in range(N_PETANI)],

    # MT3 — 30% petani aktif di MT3
    'Komoditas MT3': [np.random.choice(komoditas_list) if np.random.random() > 0.7 else '' for _ in range(N_PETANI)],
    'Luas Lahan (Ha) MT3': [round(np.random.uniform(0.25, 1.5), 2) if np.random.random() > 0.7 else 0 for _ in range(N_PETANI)],
    'Pupuk Urea (Kg) MT3':        [int(np.random.choice([50, 100, 150])) if np.random.random() > 0.7 else 0 for _ in range(N_PETANI)],
    'Pupuk NPK (Kg) MT3':         [int(np.random.choice([50, 100]))      if np.random.random() > 0.7 else 0 for _ in range(N_PETANI)],
    'Pupuk NPK Formula (Kg) MT3': [int(np.random.choice([0, 50]))        if np.random.random() > 0.7 else 0 for _ in range(N_PETANI)],
    'Pupuk Organik (Kg) MT3':     [int(np.random.choice([0, 100]))       if np.random.random() > 0.7 else 0 for _ in range(N_PETANI)],
    'Pupuk ZA (Kg) MT3':          [int(np.random.choice([0, 50]))        if np.random.random() > 0.7 else 0 for _ in range(N_PETANI)],
}

rdkk_df = pd.DataFrame(rdkk)

# =====================================================
# 2. GENERATE DATA SIVERVAL — SEMUA NORMAL
#
# Kunci agar label = NORMAL:
#   - tebus = total MT1 + MT2 + MT3 (persis, tidak dibulatkan)
#   - kios_tebus = kios_rdkk (sama persis)
#   - sp36 = 0, organik_cair = 0
#   - TIDAK ada blok noise yang mengubah nilai tebus
# =====================================================
siverval_records = []

for i in range(N_PETANI):
    nik  = niks[i]
    nama = nama_petani[i]
    kios = kode_kios_rdkk[i]

    # Total diajukan = MT1 + MT2 + MT3 (integer persis)
    urea_total   = int(rdkk['Pupuk Urea (Kg) MT1'][i])        + int(rdkk['Pupuk Urea (Kg) MT2'][i])        + int(rdkk['Pupuk Urea (Kg) MT3'][i])
    npk_total    = int(rdkk['Pupuk NPK (Kg) MT1'][i])         + int(rdkk['Pupuk NPK (Kg) MT2'][i])         + int(rdkk['Pupuk NPK (Kg) MT3'][i])
    za_total     = int(rdkk['Pupuk ZA (Kg) MT1'][i])          + int(rdkk['Pupuk ZA (Kg) MT2'][i])          + int(rdkk['Pupuk ZA (Kg) MT3'][i])
    npkf_total   = int(rdkk['Pupuk NPK Formula (Kg) MT1'][i]) + int(rdkk['Pupuk NPK Formula (Kg) MT2'][i]) + int(rdkk['Pupuk NPK Formula (Kg) MT3'][i])
    org_total    = int(rdkk['Pupuk Organik (Kg) MT1'][i])     + int(rdkk['Pupuk Organik (Kg) MT2'][i])     + int(rdkk['Pupuk Organik (Kg) MT3'][i])

    siverval_records.append({
        'no':            i + 1,
        'Kabupaten':     np.random.choice(kabupaten_list),
        'kecamatan':     np.random.choice(kecamatan_list),
        'no transaksi':  f'TRX-2025-{i+1:05d}',
        'kode kios':     kios,                   # ← SAMA dengan RDKK
        'nama kios':     kios_data[kios],
        'NIK':           nik,
        'nama petani':   nama,
        'urea':          urea_total,             # ← PERSIS sama dengan diajukan
        'npk':           npk_total,
        'sp36':          0,                      # ← tidak ada SP36
        'za':            za_total,
        'npk formula':   npkf_total,
        'organik':       org_total,
        'organik cair':  0,                      # ← tidak ada organik cair
        'tgl tebus':     f'2025-{np.random.randint(1,7):02d}-{np.random.randint(1,28):02d}',
        'tgl input':     f'2025-{np.random.randint(1,7):02d}-{np.random.randint(1,28):02d}',
        'status petani': 'Aktif',
    })

siverval_df = pd.DataFrame(siverval_records)

# =====================================================
# 3. SIMPAN KE EXCEL
# Catatan: TIDAK ada blok noise — noise akan merusak selisih = 0
# =====================================================
rdkk_path     = os.path.join('dummy_data', 'data_rdkk_normal.xlsx')
siverval_path = os.path.join('dummy_data', 'data_siverval_normal.xlsx')

rdkk_df.to_excel(rdkk_path, index=False, engine='openpyxl')

wb = Workbook()
ws = wb.active
ws.title = 'Si Verval'

total_kolom = len(siverval_df.columns)
last_col_letter = chr(64 + total_kolom) if total_kolom <= 26 else 'S'
ws.merge_cells(f'A1:{last_col_letter}1')
ws['A1'] = 'Si Verval - Kementerian Pertanian'
ws['A1'].font      = Font(bold=False, size=11)
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

for col_idx, col_name in enumerate(siverval_df.columns, 1):
    cell = ws.cell(row=2, column=col_idx, value=col_name)
    cell.font      = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

for row_idx, row in enumerate(siverval_df.values, 3):
    for col_idx, value in enumerate(row, 1):
        ws.cell(row=row_idx, column=col_idx, value=value)

wb.save(siverval_path)

print('=' * 55)
print('✅ DATA DUMMY NORMAL BERHASIL DIBUAT!')
print('=' * 55)
print(f'  RDKK    : {rdkk_path}  ({len(rdkk_df)} petani)')
print(f'  SIVERVAL: {siverval_path}  ({len(siverval_df)} transaksi)')
print()
print('Kenapa semua NORMAL?')
print('  ✓ tebus = total MT1+MT2+MT3 persis (selisih=0)')
print('  ✓ kios tebus = kios RDKK')
print('  ✓ sp36 = 0, organik_cair = 0')
print('  ✓ tidak ada blok noise yang mengubah nilai')
print()
print('Gunakan file ini untuk menguji endpoint /api/predict')
print('setelah training dengan data campuran normal.')