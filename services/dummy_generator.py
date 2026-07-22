import io
import os
from datetime import date
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

KECAMATAN_DESA_FALLBACK = {
    'Kecamatan Serang': [
        ('3673011001', 'SERANG'), ('3673011002', 'CIPARE'), ('3673011003', 'SUMUR PECUNG'),
        ('3673011004', 'KOTA BARU'), ('3673011005', 'LOPANG'), ('3673011006', 'CIMUNCANG'),
        ('3673011007', 'UNYUR'), ('3673011008', 'SUKAWANA'), ('3673011009', 'LONTARBARU'),
        ('3673011010', 'KALIGANDU'), ('3673011011', 'TERONDOL'), ('3673011012', 'KAGUNGAN'),
    ],
    'Kecamatan Kasemen': [
        ('3673021001', 'KASEMEN'), ('3673021002', 'MESJID PRIYAYI'), ('3673021003', 'TERUMBU'),
        ('3673021004', 'WARUNG JAUD'), ('3673021006', 'BENDUNG'), ('3673021007', 'BANTEN'),
        ('3673021008', 'SAWAH LUHUR'), ('3673021009', 'KILASAH'), ('3673021010', 'KASUNYATAN'),
        ('3673021011', 'MARGALUYU'),
    ],
    'Kecamatan Walantaka': [
        ('3673031001', 'WALANTAKA'), ('3673031002', 'CIGOONG'), ('3673031003', 'NYAPAH'),
        ('3673031004', 'PANGAMPELAN'), ('3673031006', 'KIARA'), ('3673031007', 'PAGER AGUNG'),
        ('3673031008', 'KALODRAN'), ('3673031009', 'KAPUREN'), ('3673031010', 'TERITIH'),
        ('3673031011', 'PABUARAN'), ('3673031012', 'PASULUHAN'), ('3673031013', 'TEGALSARI'),
        ('3673031014', 'PIPITAN'), ('3673031016', 'LEBAKWANGI'),
    ],
    'Kecamatan Curug': [
        ('3673041001', 'CURUG'), ('3673041002', 'TINGGAR'), ('3673041003', 'KEMANISAN'),
        ('3673041004', 'CIPETE'), ('3673041005', 'CILAKU'), ('3673041006', 'PANCALAKSANA'),
        ('3673041007', 'SUKAWANA'), ('3673041008', 'SUKALAKSANA'), ('3673041009', 'CURUG MANIS'),
        ('3673041010', 'SUKAJAYA'),
    ],
    'Kecamatan Cipocok Jaya': [
        ('3673051001', 'CIPOCOK JAYA'), ('3673051002', 'KARUNDANG'), ('3673051003', 'PANANCANGAN'),
        ('3673051004', 'BANJAR AGUNG'), ('3673051005', 'BANJARSARI'), ('3673051006', 'TEMBONG'),
        ('3673051007', 'DALUNG'), ('3673051008', 'GELAM'),
    ],
    'Kecamatan Taktakan': [
        ('3673061001', 'TAKTAKAN'), ('3673061002', 'SAYAR'), ('3673061003', 'PANCUR'),
        ('3673061004', 'KURANJI'), ('3673061005', 'KALANGANYAR'), ('3673061006', 'CILOWONG'),
        ('3673061007', 'PANGGUNGJATI'), ('3673061008', 'DRANGONG'), ('3673061009', 'UMBUL TENGAH'),
        ('3673061010', 'SEPANG'), ('3673061011', 'LIALANG'), ('3673061012', 'TAMAN BARU'),
    ],
}

_kecamatan_desa_cache: dict | None = None

KECAMATAN_WEIGHTS = {
    'Kasemen': 0.420,
    'Walantaka': 0.230,
    'Taktakan': 0.119,
    'Curug': 0.109,
    'Cipocok Jaya': 0.089,
    'Serang': 0.033,
}

KECAMATAN_NAMES = list(KECAMATAN_WEIGHTS.keys())
KECAMATAN_PROBS = list(KECAMATAN_WEIGHTS.values())

SUBSEKTOR_KOMODITAS = {
    'HORTIKULTURA': ['CABAI'],
    'TANAMAN PANGAN': ['PADI', 'JAGUNG', 'UBI KAYU'],
}

KECAMATAN_LAHAN = {
    'Kasemen':      (0.15, 2.00, 0.50),
    'Walantaka':    (0.20, 2.00, 0.60),
    'Taktakan':     (0.07, 1.50, 0.30),
    'Curug':        (0.10, 1.80, 0.40),
    'Cipocok Jaya': (0.07, 1.20, 0.30),
    'Serang':       (0.07, 1.00, 0.25),
}

KECAMATAN_SUPSEKTOR = {
    'Kasemen':      [('TANAMAN PANGAN', 0.90), ('HORTIKULTURA', 0.10)],
    'Walantaka':    [('TANAMAN PANGAN', 0.95), ('HORTIKULTURA', 0.05)],
    'Taktakan':     [('TANAMAN PANGAN', 0.60), ('HORTIKULTURA', 0.40)],
    'Curug':        [('TANAMAN PANGAN', 0.80), ('HORTIKULTURA', 0.20)],
    'Cipocok Jaya': [('TANAMAN PANGAN', 0.70), ('HORTIKULTURA', 0.30)],
    'Serang':       [('TANAMAN PANGAN', 0.65), ('HORTIKULTURA', 0.35)],
}

DOSIS_PER_HA = {
    'PADI':     {'urea': (200, 300), 'npk': (200, 300)},
    'CABAI':    {'urea': (100, 200), 'npk': (250, 500)},
    'JAGUNG':   {'urea': (200, 300), 'npk': (200, 300)},
    'UBI KAYU': {'urea': (100, 150), 'npk': (150, 250)},
}

NAMA_PRIA = [
    'Asep', 'Ahmad', 'Budi', 'Cecep', 'Dede', 'Eko', 'Fajar', 'Hendra',
    'Indra', 'Joko', 'Maman', 'Nana', 'Rahmat', 'Slamet', 'Taufik',
    'Ujang', 'Wahyu', 'Yana', 'Zainal', 'Acep', 'Ade', 'Agus', 'Bambang',
    'Cucu', 'Dani', 'Dedi', 'Entis', 'Fadli', 'Guntur', 'Hadi', 'Iwan',
    'Karta', 'Mulyadi', 'Mulyana', 'Oki', 'Rizky', 'Rudi', 'Saepulloh',
    'Soleh', 'Teten', 'Usep', 'Yanto', 'Yusuf', 'Zaki', 'Jajang',
    'Ating', 'Wawan', 'Jaja',
]

NAMA_WANITA = [
    'Ani', 'Dewi', 'Eneng', 'Euis', 'Fitria', 'Gita', 'Iis', 'Kartika',
    'Lestari', 'Mega', 'Nining', 'Nurma', 'Putri', 'Rina', 'Sari',
    'Siti', 'Sri', 'Titin', 'Wati', 'Yanti', 'Yayah', 'Cucu', 'Nenden',
    'Tati', 'Imas', 'Yeti', 'Elis', 'Ai', 'Nina', 'Yani',
]

PENYULUH_LIST = [
    'PUTRI KEMALA DEWI, S.P', 'DR. IR. H. AHMAD SUHENDAR, M.P',
    'SITI NURJANAH, S.P', 'BAMBANG Setiawan, S.Pt',
    'RINA WIDIASTUTI, S.P', 'EKO PRASETYO, S.P',
]

KECAMATAN_POKTAN = {
    'Kasemen': [
        'PRIANGAN MUKTI I', 'HARJA MUKTI II', 'SRI ASIH', 'SUMBER REZEKI',
        'GUYUB I', 'TUNAS ALAM', 'PUTRA MANDIRI', 'BINA SEJAHTERA',
        'SURYA JAYA', 'PERSADA', 'SUMBER JAYA',
    ],
    'Walantaka': [
        'ALAM MAKMUR', 'MAKMUR III', 'TANI GUYUB', 'MARGAHAYU 1',
        'BSR', 'KUBANG WARA',
    ],
    'Taktakan': [
        'MAKMUR JAYA II', 'GAPOKTAN RUKUN TANI', 'KWT SRI REJEKI', 'GONDANG JAYA',
    ],
    'Curug': [
        'SAUYUNAN 1', 'HARJA TANI', 'JAGABAYA MANDIRI',
    ],
    'Cipocok Jaya': [
        'JABON ADIASA MANDIRI', 'KTH KARUNDANG',
    ],
    'Serang': [
        'IKHLAS TANI SEJAHTERA (ITS)', 'PUTRA BAKTI MULYA',
    ],
}
POKTAN_FALLBACK = ['POKTAN SETEMPAT']

SUPSEKTOR_LIST = ['HORTIKULTURA', 'TANAMAN PANGAN']

KIOS_PER_KECAMATAN = {
    'Serang': [
        ('RT0000003780', 'PUSAKA TANI I'),
        ('RT0000003781', 'BINTANG TANI'),
        ('RT0000003782', 'RATA JAYA 2'),
        ('RT0000003783', 'MULTI HARAPAN TANI'),
    ],
    'Kasemen': [
        ('RT0000003784', 'TANI MULYA'),
        ('RT0000003785', 'BERKAH TANI'),
        ('RT0000003786', 'MARGALUYU UTAMA'),
        ('RT0000003787', 'TANI SUBUR'),
    ],
    'Walantaka': [
        ('RT0000003788', 'SEDERHANA'),
        ('RT0000003789', 'WALANTAKA TANI'),
        ('RT0000003790', 'TRI JAYA'),
    ],
    'Curug': [
        ('RT0000003791', 'CURUG TANI MAKMUR'),
        ('RT0000003792', 'BERKAH MANDIRI'),
    ],
    'Taktakan': [
        ('RT0000003793', 'DRANGONG'),
    ],
    'Cipocok Jaya': [
        ('RT0000003794', 'CIPOCOK TANI'),
    ],
}

TEMPAT_LAHIR = ['SERANG', 'LEBAK', 'CILEGON', 'PANDEGLANG', 'TANGERANG', 'JAKARTA', 'BANDUNG', 'LAMPUNG']

ALAMAT_TEMPLATE = {
    'Serang': [
        'KP. {desa} RT{rnd1:02d}/{rnd2:02d}',
        'KP. {desa} RT{rnd1:02d}/{rnd2:02d} KEL. {desa} KEC. SERANG KOTA SERANG',
        'CIRACAS INDAH BLOK {blok} NO {rnd3} RT{rnd1:02d}/12 SERANG',
        'PERUMNAS CIRACAS INDAH BLOK {blok} NO {rnd3} RT{rnd1:02d}/{rnd2:02d} SERANG',
        'KP. {desa} RT{rnd1:02d}/{rnd2:02d} KOTA SERANG',
        'LINK. CIRACAS III BLOK {blok} RT{rnd1:02d}/{rnd2:02d} KEL. SERANG',
    ],
    'Kasemen': [
        'KP. {desa} RT{rnd1:02d}/{rnd2:02d}',
        'KP. {desa} RT{rnd1:02d}/{rnd2:02d} KEL. {desa} KEC. KASEMEN KOTA SERANG',
        'JL. RAYA KASEMEN KM {rnd3} RT{rnd1:02d}/{rnd2:02d}',
        'KP. PESISIR BANTEN RT{rnd1:02d}/{rnd2:02d} KEC. KASEMEN',
        'PERUMNAS BANTEN INDAH BLOK {blok} NO {rnd3} RT{rnd1:02d}/{rnd2:02d}',
    ],
    'Walantaka': [
        'KP. {desa} RT{rnd1:02d}/{rnd2:02d}',
        'KP. {desa} RT{rnd1:02d}/{rnd2:02d} KEL. {desa} KEC. WALANTAKA KOTA SERANG',
        'JL. RAYA WALANTAKA KM {rnd3} KP. {desa}',
        'KP. PAGERAGUNG RT{rnd1:02d}/{rnd2:02d} KEC. WALANTAKA',
        'KP. TERITIH GIRANG RT{rnd1:02d}/{rnd2:02d}',
    ],
    'Curug': [
        'KP. {desa} RT{rnd1:02d}/{rnd2:02d}',
        'KP. {desa} RT{rnd1:02d}/{rnd2:02d} KEL. {desa} KEC. CURUG KOTA SERANG',
        'JL. RAYA CURUG KM {rnd3} KP. {desa}',
        'KP. CURUG MANIS RT{rnd1:02d}/{rnd2:02d} KEC. CURUG',
        'PERUMNAS CURUG INDAH BLOK {blok} NO {rnd3}',
    ],
    'Cipocok Jaya': [
        'KP. {desa} RT{rnd1:02d}/{rnd2:02d}',
        'KP. {desa} RT{rnd1:02d}/{rnd2:02d} KEL. {desa} KEC. CIPOCOK JAYA KOTA SERANG',
        'PERUM BANJARSARI BLOK {blok} NO {rnd3} RT{rnd1:02d}/{rnd2:02d}',
        'PERUMNAS CIPOCOK INDAH BLOK {blok} RT{rnd1:02d}/{rnd2:02d}',
        'KP. GELAM RT{rnd1:02d}/{rnd2:02d} KEC. CIPOCOK JAYA',
    ],
    'Taktakan': [
        'KP. {desa} RT{rnd1:02d}/{rnd2:02d}',
        'KP. {desa} RT{rnd1:02d}/{rnd2:02d} KEL. {desa} KEC. TAKTAKAN KOTA SERANG',
        'JL. RAYA DRANGONG KM {rnd3} KP. {desa}',
        'KP. CILOWONG RT{rnd1:02d}/{rnd2:02d} KEC. TAKTAKAN',
        'KP. PANGGUNGJATI RT{rnd1:02d}/{rnd2:02d}',
    ],
}

ALAMAT_LUAR = [
    'KP. KARANG JAYA RT 09 KARANG MARITIM PANJANG LAMPUNG',
    'KP. CIBANDUNG RT 07/10',
    'KP. PASAR KEMIS RT 03/05 TANGERANG',
    'KP. LEUWILIANG RT 12/04 BOGOR',
    'KP. BABAKAN RT 05/08 CIREBON',
    'KP. SINDANGKASIH RT 04/06 CIAMIS',
    'KP. MEKAR JAYA RT 11/02 LAMPUNG TIMUR',
]


def _load_kecamatan_desa() -> dict:
    global _kecamatan_desa_cache
    if _kecamatan_desa_cache is not None:
        return _kecamatan_desa_cache

    try:
        from config.supabase_client import get_supabase
        supabase = get_supabase()
        result = supabase.table('kecamatan_desa').select('kode_desa, nama_desa, kecamatan').execute()
        if result.data:
            data: dict = {}
            for row in result.data:
                key = f"Kecamatan {row['kecamatan'].strip().title()}"
                if key not in data:
                    data[key] = []
                data[key].append((row['kode_desa'], row['nama_desa']))
            _kecamatan_desa_cache = data
            return data
    except Exception:
        pass

    _kecamatan_desa_cache = KECAMATAN_DESA_FALLBACK
    return _kecamatan_desa_cache


def _generate_nik(rng, tgl_lahir: str, is_wanita: bool, used_niks: set) -> str:
    year, month, day = tgl_lahir.split('-')
    dd = int(day)
    mm = month
    yy = year[2:4]

    if is_wanita:
        dd += 40
    dd_str = f'{dd:02d}'

    if rng.random() < 0.8:
        kode_kec = rng.choice(['01', '02', '03', '04', '05', '06'])
        prefix = f'3673{kode_kec}'
    elif rng.random() < 0.5:
        kode_kec = rng.choice(['01', '02', '03', '04', '05', '06', '07', '08', '09', '10', '11', '12', '13', '14', '15'])
        prefix = f'3604{kode_kec}'
    else:
        prefix = rng.choice(['367101', '367201', '360101', '360201', '360301'])

    ddmm = f'{dd_str}{mm}'
    for _ in range(10000):
        suffix = f'{rng.integers(0, 10000):04d}'
        nik = f'`{prefix}{ddmm}{yy}{suffix}'
        if nik not in used_niks:
            used_niks.add(nik)
            return nik

    nik = f'`{prefix}{ddmm}{yy}{rng.integers(0, 10000):04d}'
    used_niks.add(nik)
    return nik


def _get_desa(kecamatan_filter: str | None, kecamatan_desa_data: dict):
    if kecamatan_filter:
        desa_list = kecamatan_desa_data.get(kecamatan_filter, [])
        if not desa_list:
            raise ValueError(f"Kecamatan '{kecamatan_filter}' tidak ditemukan")
        return desa_list
    all_desa = []
    for desas in kecamatan_desa_data.values():
        all_desa.extend(desas)
    return all_desa


def _generate_nama(rng, is_pria: bool) -> str:
    pool = NAMA_PRIA if is_pria else NAMA_WANITA
    if rng.random() < 0.3:
        return str(rng.choice(pool))
    return f'{rng.choice(pool)} {rng.choice(pool)}'


def _generate_ibu_kandung(rng) -> str:
    if rng.random() < 0.87:
        return str(rng.choice(NAMA_WANITA))
    return f'{rng.choice(NAMA_WANITA)} {rng.choice(NAMA_WANITA)}'


def _generate_alamat(rng, nama_desa: str, kecamatan: str) -> str:
    if rng.random() < 0.1:
        return str(rng.choice(ALAMAT_LUAR))

    templates = ALAMAT_TEMPLATE.get(kecamatan, ALAMAT_TEMPLATE['Serang'])
    template = rng.choice(templates)
    blok = rng.choice(['A', 'B', 'C', 'D', 'E', 'F'])
    return template.format(
        desa=nama_desa,
        rnd1=int(rng.integers(1, 14)),
        rnd2=int(rng.integers(1, 18)),
        rnd3=int(rng.integers(1, 20)),
        blok=blok,
    )


def _calc_pupuk(rng, komoditas: str, luas: float) -> tuple:
    if luas <= 0 or komoditas == '-':
        return (0, 0)
    dosis = DOSIS_PER_HA.get(komoditas, DOSIS_PER_HA['PADI'])
    urea = int(luas * rng.uniform(*dosis['urea']))
    npk = int(luas * rng.uniform(*dosis['npk']))
    urea = max(0, urea)
    npk = max(0, npk)
    return (urea, npk)


def _kg_val(val):
    if val == '-' or val is None:
        return 0
    if isinstance(val, float) and (pd.isna(val) or np.isnan(val)):
        return 0
    return int(val)


def generate_dummy_data(
    n_petani: int = 350,
    n_transaksi: int = 260,
    seed: int | None = None,
    pct_normal: float = 50,
    pct_over: float = 8,
    pct_kurang: float = 42,
    kecamatan_filter: str | None = None,
):
    if abs(pct_normal + pct_over + pct_kurang - 100) > 0.01:
        raise ValueError('Persentase skenario harus berjumlah 100%')

    rng = np.random.default_rng(seed)
    seed_val = seed if seed is not None else int(rng.integers(0, 2**31))

    kecamatan_desa_data = _load_kecamatan_desa()
    all_desa = _get_desa(kecamatan_filter, kecamatan_desa_data)

    kec_names = []
    desa_by_kec = {}
    for kec_key, desas in kecamatan_desa_data.items():
        kn = kec_key.replace('Kecamatan ', '')
        kec_names.append(kn)
        desa_by_kec[kn] = desas

    petani_kode_desa = []
    petani_nama_desa = []
    petani_kecamatan = []
    petani_tgl_lahir = []
    petani_nik = []
    petani_nama = []
    petani_ibu = []
    petani_alamat = []
    petani_tempat_lahir = []
    petani_komoditas_mt1 = []
    petani_luas_mt1 = []
    petani_urea_mt1 = []
    petani_npk_mt1 = []
    petani_komoditas_mt2 = []
    petani_luas_mt2 = []
    petani_urea_mt2 = []
    petani_npk_mt2 = []
    petani_subsektor = []
    petani_kode_kios = []
    petani_nama_kios = []
    used_niks: set = set()
    dash_row = ['-'] * n_petani
    na_row = [pd.NA] * n_petani

    for i in range(n_petani):
        if kecamatan_filter:
            kec = kecamatan_filter.replace('Kecamatan ', '')
            desa_list = kecamatan_desa_data.get(kecamatan_filter, [])
        else:
            kec = rng.choice(KECAMATAN_NAMES, p=KECAMATAN_PROBS)
            desa_list = desa_by_kec[kec]

        petani_kecamatan.append(kec)

        kios_list = KIOS_PER_KECAMATAN.get(kec, KIOS_PER_KECAMATAN['Serang'])
        kk, nk = kios_list[rng.integers(0, len(kios_list))]
        petani_kode_kios.append(kk)
        petani_nama_kios.append(nk)

        kd, nd = desa_list[rng.integers(0, len(desa_list))]
        petani_kode_desa.append(kd)
        petani_nama_desa.append(nd)

        tgl = f'{rng.integers(1960, 1995)}-{rng.integers(1, 13):02d}-{rng.integers(1, 29):02d}'
        petani_tgl_lahir.append(tgl)

        petani_tempat_lahir.append(str(rng.choice(TEMPAT_LAHIR)))

        is_wanita = rng.random() >= 0.7
        petani_nik.append(_generate_nik(rng, tgl, is_wanita, used_niks))

        petani_nama.append(_generate_nama(rng, not is_wanita).upper())
        petani_ibu.append(_generate_ibu_kandung(rng).upper())
        petani_alamat.append(_generate_alamat(rng, nd, kec))

        sub_choices = KECAMATAN_SUPSEKTOR.get(kec, SUPSEKTOR_LIST)
        sub_items, sub_probs = zip(*sub_choices)
        subsektor = str(rng.choice(sub_items, p=sub_probs))
        petani_subsektor.append(subsektor)

        kom_items = SUBSEKTOR_KOMODITAS[subsektor]
        kom_mt1 = str(rng.choice(kom_items))
        petani_komoditas_mt1.append(kom_mt1)

        lahan_params = KECAMATAN_LAHAN.get(kec, (0.1, 2.0, 0.5))
        low, high, mode = lahan_params
        luas_mt1 = min(rng.triangular(low, mode, high), 2.0)
        luas_mt1 = round(luas_mt1, 2)
        petani_luas_mt1.append(luas_mt1)

        urea_mt1, npk_mt1 = _calc_pupuk(rng, kom_mt1, luas_mt1)
        petani_urea_mt1.append(urea_mt1)
        petani_npk_mt1.append(npk_mt1)

        has_mt2 = rng.random() < 0.90
        if has_mt2:
            kom_items_mt2 = SUBSEKTOR_KOMODITAS[subsektor]
            kom_mt2 = str(rng.choice(kom_items_mt2))
            petani_komoditas_mt2.append(kom_mt2)

            low2, high2, mode2 = KECAMATAN_LAHAN.get(kec, (0.1, 2.0, 0.5))
            luas_mt2 = min(rng.triangular(low2, mode2, high2), 2.0)
            luas_mt2 = round(luas_mt2, 2)
            petani_luas_mt2.append(luas_mt2)

            urea_mt2, npk_mt2 = _calc_pupuk(rng, kom_mt2, luas_mt2)
            petani_urea_mt2.append(urea_mt2)
            petani_npk_mt2.append(npk_mt2)
        else:
            petani_komoditas_mt2.append('-')
            petani_luas_mt2.append(0)
            petani_urea_mt2.append(0)
            petani_npk_mt2.append(0)

    rdkk = {
        'Nama Penyuluh': [str(rng.choice(PENYULUH_LIST)) for _ in range(n_petani)],
        'Kode Desa': petani_kode_desa,
        'Kode Kios Pengecer': petani_kode_kios,
        'Nama Kios Pengecer': petani_nama_kios,
        'Gapoktan': na_row[:],
        'Nama Poktan': [str(rng.choice(KECAMATAN_POKTAN.get(kec, POKTAN_FALLBACK))) for kec in petani_kecamatan],
        'Nama Petani': petani_nama,
        'KTP': petani_nik,
        'Tempat Lahir': petani_tempat_lahir,
        'Tanggal Lahir': petani_tgl_lahir,
        'Nama Ibu Kandung': petani_ibu,
        'Alamat': petani_alamat,
        'Subsektor': petani_subsektor,
        'Komoditas MT1': petani_komoditas_mt1,
        'Luas Lahan (Ha) MT1': petani_luas_mt1,
        'Pupuk Urea (Kg) MT1': petani_urea_mt1,
        'Pupuk NPK (Kg) MT1': petani_npk_mt1,
        'Pupuk NPK Formula (Kg) MT1': dash_row[:],
        'Pupuk Organik (Kg) MT1': dash_row[:],
        'Pupuk ZA (Kg) MT1': dash_row[:],
        'Komoditas MT2': petani_komoditas_mt2,
        'Luas Lahan (Ha) MT2': petani_luas_mt2,
        'Pupuk Urea (Kg) MT2': petani_urea_mt2,
        'Pupuk NPK (Kg) MT2': petani_npk_mt2,
        'Pupuk NPK Formula (Kg) MT2': dash_row[:],
        'Pupuk Organik (Kg) MT2': dash_row[:],
        'Pupuk ZA (Kg) MT2': dash_row[:],
        'Komoditas MT3': dash_row[:],
        'Luas Lahan (Ha) MT3': dash_row[:],
        'Pupuk Urea (Kg) MT3': dash_row[:],
        'Pupuk NPK (Kg) MT3': dash_row[:],
        'Pupuk NPK Formula (Kg) MT3': dash_row[:],
        'Pupuk Organik (Kg) MT3': dash_row[:],
        'Pupuk ZA (Kg) MT3': dash_row[:],
    }

    rdkk_df = pd.DataFrame(rdkk)

    weights = np.array([pct_normal, pct_over, pct_kurang]) / 100.0
    farmer_scenario = rng.choice(3, size=n_petani, p=weights).tolist()
    scenario_map = {0: 'normal', 1: 'over', 2: 'kurang'}

    potential_trans = []
    for i in range(n_petani):
        potential_trans.append((i, 'MT1'))
        if petani_komoditas_mt2[i] != '-':
            potential_trans.append((i, 'MT2'))

    n_pool = len(potential_trans)
    if n_transaksi <= n_pool:
        selected = rng.choice(range(n_pool), n_transaksi, replace=False).tolist()
    else:
        selected = list(range(n_pool))
        extra = rng.choice(range(n_pool), n_transaksi - n_pool, replace=True).tolist()
        selected.extend(extra)

    siverval_records = []
    for no, pool_idx in enumerate(selected, 1):
        farmer_idx, mt = potential_trans[pool_idx]
        sc_key = farmer_scenario[farmer_idx]
        scenario = scenario_map[sc_key]
        nik_rdkk = petani_nik[farmer_idx]
        nik_sv = "'" + nik_rdkk[1:]
        nama = petani_nama[farmer_idx]

        if mt == 'MT1':
            urea_total = petani_urea_mt1[farmer_idx]
            npk_total = petani_npk_mt1[farmer_idx]
            komoditas = petani_komoditas_mt1[farmer_idx]
        else:
            urea_total = petani_urea_mt2[farmer_idx]
            npk_total = petani_npk_mt2[farmer_idx]
            komoditas = petani_komoditas_mt2[farmer_idx]

        if scenario == 'normal':
            urea_tebus = urea_total
            npk_tebus = npk_total
        elif scenario == 'over':
            urea_tebus = int(urea_total * rng.uniform(1.1, 2.0))
            npk_tebus = int(npk_total * rng.uniform(1.0, 1.5))
        elif scenario == 'kurang':
            urea_tebus = int(urea_total * rng.uniform(0.1, 0.84))
            npk_tebus = int(npk_total * rng.uniform(0.1, 0.84))

        status = 'Disetujui Tim Verval Pusat'
        no_transaksi = f'{petani_kode_kios[farmer_idx]}\\{rng.choice(["V", "W"])}00{rng.choice(["X", "Y", "Z", "U"])}{chr(86 + no % 26)}'
        tgl_tebus = f'{rng.integers(1, 29)}-{rng.integers(1, 7)}-2026'
        tgl_input = f'2026-{rng.integers(1, 7):02d}-{rng.integers(1, 29):02d} {rng.integers(8, 18):02d}:{rng.integers(0, 60):02d}:00'

        siverval_records.append({
            'NO': no,
            'KABUPATEN': 'KOTA SERANG',
            'KECAMATAN': petani_kecamatan[farmer_idx].upper() if farmer_idx < len(petani_kecamatan) else 'SERANG',
            'NO TRANSAKSI': no_transaksi,
            'KODE KIOS': petani_kode_kios[farmer_idx],
            'NAMA KIOS': petani_nama_kios[farmer_idx],
            'NIK': nik_sv,
            'NAMA PETANI': nama,
            'UREA': urea_tebus,
            'NPK': npk_tebus,
            'SP36': 0,
            'ZA': 0,
            'NPK FORMULA': 0,
            'ORGANIK': 0,
            'ORGANIK CAIR': 0,
            'TGL TEBUS': tgl_tebus,
            'TGL INPUT': tgl_input,
            'STATUS': status,
        })

    siverval_df = pd.DataFrame(siverval_records)

    rdkk_buf = io.BytesIO()
    rdkk_df.to_excel(rdkk_buf, index=False, engine='openpyxl')
    rdkk_buf.seek(0)

    siverval_buf = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = 'Si Verval'

    total_kolom = len(siverval_df.columns)
    last_col = chr(64 + total_kolom) if total_kolom <= 26 else 'S'
    ws.merge_cells(f'A1:{last_col}1')
    ws['A1'] = 'Si Verval - Kementrian Pertanian'
    ws['A1'].font = Font(bold=False, size=11)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    for col_idx, col_name in enumerate(siverval_df.columns, 1):
        cell = ws.cell(row=2, column=col_idx, value=col_name)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    for row_idx, row in enumerate(siverval_df.values, 3):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    wb.save(siverval_buf)
    siverval_buf.seek(0)

    label_counts = pd.Series(farmer_scenario).map(scenario_map).value_counts()
    trans_by_scenario = pd.Series([scenario_map[farmer_scenario[i]] for i, _ in potential_trans]).value_counts()
    summary = {
        'n_petani': n_petani,
        'n_transaksi': len(selected),
        'seed': int(seed_val),
        'kecamatan': kecamatan_filter or 'Semua',
        'distribusi_skenario': {str(k): int(v) for k, v in label_counts.to_dict().items()},
    }

    return rdkk_buf, siverval_buf, summary
