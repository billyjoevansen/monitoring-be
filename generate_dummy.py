"""
Script untuk generate data dummy RDKK dan SIVERVAL.
Format meniru file aktual Rdkk.xlsx dan SIVerval.xlsx.
Bisa dijalankan berulang: python generate_dummy.py
File akan otomatis bernama: data_rdkk.xlsx, data_rdkk_1.xlsx, data_rdkk_2.xlsx, dst.
"""
import pandas as pd
import numpy as np
import os
import glob
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from dotenv import load_dotenv
load_dotenv()

N_PETANI = 350 # Jumlah pengajuan petani dummy yang di-generate
N_TRANSAKSI = 346  # Jumlah transaksi dummy yang di-generate


def fetch_kecamatan_desa():
    """Ambil data kecamatan & desa dari Supabase."""
    try:
        from config.supabase_client import get_supabase
        supabase = get_supabase()
        result = supabase.table('kecamatan_desa').select('kode_desa, nama_desa, kecamatan').execute()
        rows = result.data or []
        if not rows:
            raise ValueError("Tabel kecamatan_desa kosong")

        mapping = {}
        for row in rows:
            kec = str(row['kecamatan']).strip().title()
            mapping.setdefault(kec, []).append((str(row['kode_desa']).strip(), str(row['nama_desa']).strip()))
        print(f"  [OK] Ambil {len(rows)} desa dari Supabase ({len(mapping)} kecamatan)")
        return mapping
    except Exception as e:
        print(f"  [WARN] Gagal ambil dari Supabase: {e}")
        print("  -> Pakai data hardcoded fallback")
        return {
            'WALANTAKA': [('3673011001', 'WALANTAKA'), ('3673011002', 'CIGOONG'), ('3673011003', 'KALODRAN')],
            'CURUG': [('3673011004', 'CURUG'), ('3673011005', 'CIPETE'), ('3673011006', 'CURUG MANIS')],
            'CIPOCOK JAYA': [('3673011007', 'BANJARSARI'), ('3673011008', 'DALUNG'), ('3673011009', 'GELAM')],
            'KASEMEN': [('3673011010', 'KASEMEN'), ('3673011011', 'BANTEN'), ('3673011012', 'BENDUNG')],
            'TAKTAKAN': [('3673011013', 'TAKTAKAN'), ('3673011014', 'CILOWONG'), ('3673011015', 'DRANGONG')],
            'SERANG': [('3673011016', 'SERANG'), ('3673011017', 'CIPARE'), ('3673011018', 'KOTA BARU')],
        }


def get_next_file_number(folder: str) -> str:
    existing = glob.glob(os.path.join(folder, 'data_rdkk*.xlsx'))
    if not existing:
        return ''
    max_num = -1
    for filepath in existing:
        filename = os.path.basename(filepath)
        if filename == 'data_rdkk.xlsx':
            max_num = max(max_num, 0)
        elif filename.startswith('data_rdkk_'):
            try:
                num = int(filename.replace('data_rdkk_', '').replace('.xlsx', ''))
                max_num = max(max_num, num)
            except ValueError:
                continue
    next_num = max_num + 1
    if next_num == 0:
        return ''
    return f'_{next_num}'


def generate_dummy():
    os.makedirs('dummy_data', exist_ok=True)
    suffix = get_next_file_number('dummy_data')
    seed = abs(hash(suffix)) % (2**31)
    np.random.seed(seed)

    print("[*] Mengambil data kecamatan/desa dari Supabase...")
    kecamatan_desa_map = fetch_kecamatan_desa()
    all_desa = [(kd, nd, kec) for kec, items in kecamatan_desa_map.items() for kd, nd in items]

    # =====================================================
    # 1. GENERATE DATA RDKK (mirip Rdkk.xlsx)
    # =====================================================

    # NIK 16 digit dengan backtick prefix (seperti aktual)
    niks = [f'`36{"".join([str(np.random.randint(0, 10)) for _ in range(14)])}' for _ in range(N_PETANI)]

    # Nama-nama dummy
    nama_depan = [
        'Budi', 'Siti', 'Ahmad', 'Dewi', 'Agus', 'Sri', 'Hadi', 'Rina', 'Joko', 'Ani',
        'Wahyu', 'Lestari', 'Bambang', 'Nurma', 'Eko', 'Yanti', 'Dani', 'Wati', 'Rudi', 'Mega',
        'Andi', 'Asep', 'Cecep', 'Dedi', 'Eneng', 'Fajar', 'Gita', 'Hendra', 'Indra', 'Iwan',
        'Kartika', 'Mulyadi', 'Nana', 'Oki', 'Putri', 'Rahmat', 'Soleh', 'Taufik', 'Ujang', 'Yanto',
        'Zaki', 'Fitria', 'Rizky', 'Dewantara', 'Sari', 'Wibowo', 'Yusuf', 'Zainal', 'Fadli', 'Guntur',
        'Halim', 'Irfan', 'Jumadi', 'Kusnadi', 'Lukman', 'Mardiana', 'Slamet', 'Yuli', 'Edi', 'Nina',
    ]
    nama_belakang = [
        'Santoso', 'Wijaya', 'Hartono', 'Susanto', 'Rahayu', 'Pratama',
        'Saputra', 'Handayani', 'Kusuma', 'Hidayat', 'Lestari', 'Cahyono',
        'Setiawan', 'Purnama', 'Nugroho', 'Wulandari', 'Suryadi', 'Utami',
        'Mulyana', 'Sudarsono', 'Gunawan', 'Suhendar', 'Heryanto', 'Fauzi',
        'Ardiansyah', 'Siregar', 'Laksana', 'Kurniawan', 'Basri', 'Mahendra',
    ]
    nama_petani = [
        f'{np.random.choice(nama_depan)} {np.random.choice(nama_belakang)}'
        for _ in range(N_PETANI)
    ]

    # Kode dan Nama Kios — single kios seperti aktual
    kode_kios = 'RT0000003780'
    nama_kios = 'PUSAKA TANI I'

    # Poktan (daerah Serang/Banten)
    poktan_list = [
        'Sri Rejeki', 'Tani Mulyo', 'Karya Tani', 'Maju Jaya', 'Sawit Berkah',
        'Harapan Baru', 'Mekar Sari', 'Subur Makmur', 'Berkah Tani',
        'Sinar Jaya', 'Menteng Jaya', 'Sinar Melati', 'Mina Mukti', 'Sumber Jaya',
        'Pandan Wangi', 'Setia Kawan', 'Taruna Mekar', 'Bina Tani', 'Tani Rahayu',
        'Sida Mukti', 'Maju Bersama', 'Tani Mandiri', 'Hurip', 'Mutiara Tani',
        'Cipta Karya', 'Tunas Harapan', 'Sari Bumi', 'Agro Lestari', 'Tani Sejahtera',
        'IKHLAS TANI SEJAHTERA'
    ]

    # Penyuluh dengan nama realistis
    penyuluh_list = [
        'PUTRI KEMALA DEWI, S.P', 'DR. IR. H. AHMAD SUHENDAR, M.P',
        'SITI NURJANAH, S.P', 'BAMBANG Setiawan, S.Pt',
        'RINA WIDIASTUTI, S.P', 'EKO PRASETYO, S.P',
    ]

    # Tempat lahir
    tempat_lahir_list = [
        'SERANG', 'LEBAK', 'PANJANG', 'TAMAN SARI', 'CILEGON',
        'PANDEGLANG', 'RANGKAS BITUNG', 'MERAK', 'TANGERANG',
        'BEKASI', 'BANDUNG', 'BOGOR', 'JAKARTA', 'SEMARANG',
    ]

    # Pilih kecamatan & desa
    petani_kecamatan = []
    petani_kode_desa = []
    petani_nama_desa = []
    for _ in range(N_PETANI):
        kd, nd, kec = all_desa[np.random.randint(0, len(all_desa))]
        petani_kecamatan.append(kec)
        petani_kode_desa.append(kd)
        petani_nama_desa.append(nd)

    # Komoditas — sesuai aktual (CABAI, JAGUNG, PADI, UBI KAYU)
    komoditas_list = ['CABAI', 'JAGUNG', 'PADI', 'UBI KAYU']

    # Subsektor — sesuai aktual
    subsektor_list = ['HORTIKULTURA', 'TANAMAN PANGAN']

    # Generate RDKK data
    rdkk_data = {
        'Nama Penyuluh': [np.random.choice(penyuluh_list) for _ in range(N_PETANI)],
        'Kode Desa': petani_kode_desa,
        'Kode Kios Pengecer': [kode_kios] * N_PETANI,
        'Nama Kios Pengecer': [nama_kios] * N_PETANI,
        'Gapoktan': [None] * N_PETANI,
        'Nama Poktan': [np.random.choice(poktan_list) for _ in range(N_PETANI)],
        'Nama Petani': nama_petani,
        'KTP': niks,
        'Tempat Lahir': [np.random.choice(tempat_lahir_list) for _ in range(N_PETANI)],
        'Tanggal Lahir': [
            f'{np.random.randint(1960, 1995)}-{np.random.randint(1, 13):02d}-{np.random.randint(1, 29):02d}'
            for _ in range(N_PETANI)
        ],
        'Nama Ibu Kandung': [
            f'{np.random.choice(nama_depan)} {np.random.choice(nama_belakang)}'
            for _ in range(N_PETANI)
        ],
        'Alamat': [
            f'PERUMNAS CIRACAS INDAH BLOK C3 NO {np.random.randint(1, 200)} RT{np.random.randint(1, 12):02d}/12 SERANG'
            for _ in range(N_PETANI)
        ],
        'Subsektor': [np.random.choice(subsektor_list) for _ in range(N_PETANI)],

        # MT1 — semua petani punya data
        'Komoditas MT1': [np.random.choice(komoditas_list) for _ in range(N_PETANI)],
        'Luas Lahan (Ha) MT1': np.round(np.random.uniform(0.1, 2.0, N_PETANI), 2),
        'Pupuk Urea (Kg) MT1': np.random.choice([50, 80, 100, 120, 150, 200, 250, 300, 400, 500], N_PETANI),
        'Pupuk NPK (Kg) MT1': np.random.choice([50, 80, 100, 120, 150, 180, 200, 250, 300, 400, 480, 600], N_PETANI),
        'Pupuk NPK Formula (Kg) MT1': [None] * N_PETANI,
        'Pupuk Organik (Kg) MT1': [None] * N_PETANI,
        'Pupuk ZA (Kg) MT1': [None] * N_PETANI,

        # MT2 — ~90% petani punya data
        'Komoditas MT2': [np.random.choice(komoditas_list) if np.random.random() < 0.90 else '' for _ in range(N_PETANI)],
        'Luas Lahan (Ha) MT2': [round(np.random.uniform(0.1, 1.5), 2) if np.random.random() < 0.90 else 0 for _ in range(N_PETANI)],
        'Pupuk Urea (Kg) MT2': [np.random.choice([50, 80, 100, 120, 150, 200]) if np.random.random() < 0.90 else None for _ in range(N_PETANI)],
        'Pupuk NPK (Kg) MT2': [np.random.choice([50, 80, 100, 120, 150, 200, 250]) if np.random.random() < 0.90 else None for _ in range(N_PETANI)],
        'Pupuk NPK Formula (Kg) MT2': [None] * N_PETANI,
        'Pupuk Organik (Kg) MT2': [None] * N_PETANI,
        'Pupuk ZA (Kg) MT2': [None] * N_PETANI,

        # MT3 — kosong semua (sesuai aktual)
        'Komoditas MT3': [''] * N_PETANI,
        'Luas Lahan (Ha) MT3': [0] * N_PETANI,
        'Pupuk Urea (Kg) MT3': [None] * N_PETANI,
        'Pupuk NPK (Kg) MT3': [None] * N_PETANI,
        'Pupuk NPK Formula (Kg) MT3': [None] * N_PETANI,
        'Pupuk Organik (Kg) MT3': [None] * N_PETANI,
        'Pupuk ZA (Kg) MT3': [None] * N_PETANI,
    }

    rdkk_df = pd.DataFrame(rdkk_data)

    # =====================================================
    # 2. GENERATE DATA SIVERVAL (mirip SIVerval.xlsx)
    # =====================================================

    # Pilih N_TRANSAKSI petani acak yang punya transaksi
    petani_transaksi_idx = np.random.choice(N_PETANI, N_TRANSAKSI, replace=False)

    # Kode transaksi realistis: S02KA1\XXXXX
    s02_codes = ['C', 'S']
    s03_codes = ['V', 'W']
    s04_codes = ['X', 'Y', 'Z', 'U']

    siverval_records = []
    no = 1

    for idx in petani_transaksi_idx:
        nik = niks[idx]
        nama = nama_petani[idx]

        # Hitung total pupuk diajukan
        urea_rdkk = (rdkk_data['Pupuk Urea (Kg) MT1'][idx] or 0) + (rdkk_data['Pupuk Urea (Kg) MT2'][idx] or 0)
        npk_rdkk = (rdkk_data['Pupuk NPK (Kg) MT1'][idx] or 0) + (rdkk_data['Pupuk NPK (Kg) MT2'][idx] or 0)

        # Default values
        sp36_tebus = 0
        organik_cair_tebus = 0
        status = 'Disetujui Tim Verval Pusat'

        # Skenario penebusan
        scenario = np.random.random()

        if scenario < 0.35:
            # NORMAL: tebus semua
            urea_tebus = urea_rdkk
            npk_tebus = npk_rdkk

        elif scenario < 0.65:
            # Tidak tebus semua (sisakan sebagian) — tetap NORMAL
            urea_tebus = int(urea_rdkk * np.random.uniform(0.3, 0.9))
            npk_tebus = int(npk_rdkk * np.random.uniform(0.4, 0.95))

        elif scenario < 0.85:
            # Tebus melebihi kuota — TIDAK NORMAL
            urea_tebus = int(urea_rdkk * np.random.uniform(1.1, 2.0))
            npk_tebus = int(npk_rdkk * np.random.uniform(1.0, 1.5))
            sp36_tebus = 0
            organik_cair_tebus = 0
            status = 'Disetujui Tim Verval Pusat'

        elif scenario < 0.95:
            # Tebus pupuk di luar RDKK (SP36/Organik Cair) — TIDAK NORMAL
            urea_tebus = int(urea_rdkk * np.random.uniform(0.8, 1.0))
            npk_tebus = int(npk_rdkk * np.random.uniform(0.8, 1.0))
            sp36_tebus = np.random.choice([50, 100, 150])
            organik_cair_tebus = np.random.choice([0, 20, 50])
            status = 'Disetujui Tim Verval Pusat'

        else:
            # Petani non-aktif tapi tebus — TIDAK NORMAL
            urea_tebus = int(urea_rdkk * np.random.uniform(0.5, 1.0))
            npk_tebus = int(npk_rdkk * np.random.uniform(0.5, 1.0))
            sp36_tebus = 0
            organik_cair_tebus = 0
            status = 'Non-aktif'

        # Generate tanggal realistis
        bulan = np.random.randint(1, 7)
        hari_tebus = np.random.randint(1, 29)
        # TGL INPUT: beberapa hari setelah TGL TEBUS (0-7 hari)
        selisih_hari = np.random.choice([0, 0, 0, 1, 1, 2, 3, 5, 7])
        jam_input = np.random.randint(8, 18)
        menit_input = np.random.randint(0, 60)

        tgl_tebus_str = f'{hari_tebus}-{bulan}-2026'
        tgl_input_str = f'2026-{bulan:02d}-{hari_tebus + selisih_hari:02d} {jam_input:02d}:{menit_input:02d}:00'

        # NO TRANSAKSI realistis: S02KA1\XXXXX
        s03 = np.random.choice(s03_codes)
        s04 = np.random.choice(s04_codes)
        no_transaksi = f'S02KA1\\{s03}00{s04}{chr(86 + no)}'

        # Status: selalu "Disetujui Tim Verval Pusat"
        status = 'Disetujui Tim Verval Pusat'

        siverval_records.append({
            'NO': no,
            'KABUPATEN': 'KOTA SERANG',
            'KECAMATAN': petani_kecamatan[idx],
            'NO TRANSAKSI': no_transaksi,
            'KODE KIOS': kode_kios,
            'NAMA KIOS': nama_kios,
            'NIK': nik,
            'NAMA PETANI': nama,
            'UREA': urea_tebus,
            'NPK': npk_tebus,
            'SP36': 0,
            'ZA': 0,
            'NPK FORMULA': 0,
            'ORGANIK': 0,
            'ORGANIK CAIR': 0,
            'TGL TEBUS': tgl_tebus_str,
            'TGL INPUT': tgl_input_str,
            'STATUS': status,
        })
        no += 1

    siverval_df = pd.DataFrame(siverval_records)

    # =====================================================
    # 3. TAMBAH NOISE SEDIKIT
    # =====================================================
    # Variasi luas lahan kecil
    for i in range(0, N_PETANI, 5):
        rdkk_df.loc[i, 'Luas Lahan (Ha) MT1'] = round(np.random.uniform(0.1, 0.3), 2)

    # =====================================================
    # 4. SIMPAN KE EXCEL
    # =====================================================
    rdkk_path = os.path.join('dummy_data', f'data_rdkk{suffix}.xlsx')
    siverval_path = os.path.join('dummy_data', f'data_siverval{suffix}.xlsx')

    # Simpan RDKK (format sederhana, header di baris 1)
    rdkk_df.to_excel(rdkk_path, index=False, engine='openpyxl')

    # Simpan SIVERVAL dengan format mirip aktual
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Si Verval"

    # Baris 1: Title
    total_kolom = len(siverval_df.columns)
    last_col_letter = chr(64 + total_kolom) if total_kolom <= 26 else 'S'
    ws.merge_cells(f'A1:{last_col_letter}1')
    ws['A1'] = 'Si Verval - Kementrian Pertanian'
    ws['A1'].font = Font(bold=False, size=11)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # Baris 2: Header UPPERCASE
    for col_idx, col_name in enumerate(siverval_df.columns, 1):
        cell = ws.cell(row=2, column=col_idx, value=col_name.upper())
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Baris 3+: Data
    for row_idx, row in enumerate(siverval_df.values, 3):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Baris terakhir: Total row
    total_row = len(siverval_df) + 3
    ws.cell(row=total_row, column=1, value='TOTAL')
    ws.cell(row=total_row, column=9, value=int(siverval_df['UREA'].sum()))
    ws.cell(row=total_row, column=10, value=int(siverval_df['NPK'].sum()))
    ws.cell(row=total_row, column=11, value=0)
    ws.cell(row=total_row, column=12, value=0)
    ws.cell(row=total_row, column=13, value=0)
    ws.cell(row=total_row, column=14, value=0)
    ws.cell(row=total_row, column=15, value=0)

    wb.save(siverval_path)

    # =====================================================
    # 5. OUTPUT
    # =====================================================
    print("=" * 60)
    print("DATA DUMMY BERHASIL DIBUAT!")
    print("=" * 60)
    print(f"\nFile tersimpan di:")
    print(f"   - {rdkk_path}")
    print(f"   - {siverval_path}")
    print(f"\nStatistik:")
    print(f"   - RDKK    : {len(rdkk_df)} petani")
    print(f"   - SIVERVAL: {len(siverval_df)} transaksi ({len(siverval_df)/len(rdkk_df)*100:.1f}%)")
    print(f"\nSkenario penebusan:")
    print(f"   - ~35% Normal (tebus semua)")
    print(f"   - ~30% Tidak tebus semua (tetap NORMAL)")
    print(f"   - ~20% Tebus melebihi kuota (TIDAK NORMAL)")
    print(f"   - ~15% Tebus di luar RDKK (TIDAK NORMAL)")
    print(f"\nFormat file:")
    print(f"   - RDKK: header baris 1, NIK dengan backtick, single kios")
    print(f"   - SIVerval: title baris 1, header UPPERCASE baris 2, TGL D-M-YYYY")

    all_files = sorted(glob.glob(os.path.join('dummy_data', '*.xlsx')))
    print(f"\nSemua file di dummy_data/:")
    for f in all_files:
        print(f"   - {os.path.basename(f)}")


if __name__ == '__main__':
    generate_dummy()
